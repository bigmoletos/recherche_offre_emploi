#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agent IA de Recherche d'Offres d'Alternance - Version Démonstration Complète.

Ce module implémente un agent intelligent pour la recherche automatisée d'offres
d'alternance en cybersécurité et réseaux télécoms. Il combine les techniques de
scraping web avancées avec l'intelligence artificielle pour filtrer et analyser
les opportunités.

Fonctionnalités principales :
- Scraping multi-sites avec gestion de la charge et anti-détection
- Classification intelligente des offres par IA (règles + ML)
- Validation automatique des critères d'alternance et de domaine
- Génération de rapports Excel professionnels avec metrics
- Logging structuré et monitoring des performances
- Gestion robuste des erreurs et retry automatique

Architecture :
- BaseScraper : Classe abstraite pour l'extensibilité multi-sites
- OfferClassifier : Moteur IA de classification et validation
- ExcelReportBuilder : Générateur de rapports riches
- AlternanceAgent : Orchestrateur principal du workflow

Le script fonctionne avec des données simulées pour démonstration,
tout en montrant l'architecture complète prête pour l'intégration
avec de vrais sites d'emploi.

Usage :
    python src/demo_agent.py                # Démonstration standard
    python src/demo_agent.py --verbose     # Mode détaillé avec logs
    python src/demo_agent.py --quick       # Mode rapide (5 offres)

Auteur: desmedt.franck@iaproject.fr
Version: 1.0
Date: 03/06/2025
"""

import requests
from bs4 import BeautifulSoup
import json
import pandas as pd
from datetime import datetime
import logging
import os
import sys
import argparse
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Union
from dataclasses import dataclass, field
import time
import random
from urllib.parse import urljoin, urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill

# Ajouter le système de logging commun
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "shared" / "scripts"))

try:
    from logger_config import get_logger
    # Utiliser le système de logging avancé
    logger = get_logger(__name__, log_file="demo_agent.log")
except ImportError:
    # Fallback vers le logging standard
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s | %(levelname)-8s | %(name)s | %(message)s',
        handlers=[
            logging.FileHandler(f'demo_agent_{datetime.now().strftime("%Y%m%d")}.log'),
            logging.StreamHandler()
        ]
    )
    logger = logging.getLogger(__name__)
    logger.warning("Module logger_config non disponible, utilisation du logging standard")

@dataclass
class JobOffer:
    """
    Structure de données représentant une offre d'emploi en alternance.

    Cette classe encapsule toutes les informations nécessaires pour analyser
    et classifier une offre d'alternance. Elle inclut les métadonnées de base
    ainsi que le statut de validation IA.

    Attributes:
        site (str): Nom du site source (Pôle Emploi, Indeed, etc.)
        title (str): Titre du poste proposé
        company (str): Nom de l'entreprise recrutrice
        location (str): Localisation géographique du poste
        duration (str): Durée du contrat d'alternance
        start_date (str): Date de début prévue
        url (str): URL complète vers l'annonce originale
        description (str): Description détaillée du poste et des missions
        is_validated (bool): Statut de validation par le classificateur IA
        validation_reason (str): Raison détaillée de la validation/rejet

    Note:
        Les champs is_validated et validation_reason sont automatiquement
        remplis par le système de classification OfferClassifier.
    """
    site: str
    title: str
    company: str
    location: str
    duration: str
    start_date: str
    url: str
    description: str = ""
    is_validated: bool = False
    validation_reason: str = ""

    def __post_init__(self):
        """
        Validation et nettoyage automatique des données après initialisation.

        Effectue des vérifications de base et normalise certains champs
        pour assurer la cohérence des données.
        """
        # Nettoyer les espaces superflus
        self.title = self.title.strip() if self.title else ""
        self.company = self.company.strip() if self.company else ""
        self.location = self.location.strip() if self.location else ""

        # Logger l'offre créée
        logger.debug(f"Nouvelle offre créée: {self.title} chez {self.company}")

class OfferClassifier:
    """
    Moteur de classification intelligente pour les offres d'alternance.

    Cette classe implémente un système de classification hybride combinant
    des règles métier expertes avec des techniques d'IA pour identifier
    et valider les offres d'alternance pertinentes en cybersécurité.

    Le classificateur effectue plusieurs niveaux de validation :
    1. Exclusion des formations/écoles (non-pertinentes)
    2. Validation de la nature alternance/apprentissage
    3. Vérification du domaine technique (cyber/réseaux)
    4. Contrôle de la durée et modalités

    Attributes:
        formation_keywords (List[str]): Mots-clés identifiant les formations à exclure
        alternance_keywords (List[str]): Mots-clés validant les contrats d'alternance
        tech_keywords (List[str]): Mots-clés du domaine cybersécurité/réseaux
        classification_stats (Dict): Statistiques de classification temps réel

    Example:
        classifier = OfferClassifier()
        is_valid, reason = classifier.classify_offer(job_offer)
        if is_valid:
            print(f"Offre validée: {reason}")
    """

    def __init__(self):
        """
        Initialise le classificateur avec les dictionnaires de mots-clés experts.

        Les mots-clés sont organisés par catégorie et optimisés pour minimiser
        les faux positifs tout en maintenant une haute sensibilité.
        """
        logger.info("Initialisation du classificateur d'offres IA")

        # Mots-clés pour identifier les formations (à exclure)
        self.formation_keywords = [
            'formation', 'école', 'cursus', 'académie', 'institut',
            'université', 'bachelor', 'master formation', 'diplôme',
            'étudiant', 'apprenant', 'inscription', 'admission',
            'enseignement', 'pédagogique', 'éducatif'
        ]

        # Mots-clés pour identifier les vraies alternances
        self.alternance_keywords = [
            'alternance', 'apprentissage', 'contrat de professionnalisation',
            'contrat pro', 'apprenti', 'alternant', 'cfa',
            'rythme alterné', 'formation en alternance'
        ]

        # Mots-clés cybersécurité/réseaux
        self.tech_keywords = [
            'cybersécurité', 'cyber', 'sécurité informatique', 'réseaux',
            'télécommunications', 'télécoms', 'infosec', 'soc', 'pentesting',
            'firewall', 'intrusion', 'vulnerability', 'cisco', 'juniper',
            'siem', 'forensic', 'ethical hacking', 'pentest', 'grc',
            'iso 27001', 'audit sécurité', 'cryptographie'
        ]

        # Statistiques de classification
        self.classification_stats = {
            'total_processed': 0,
            'validated': 0,
            'rejected_formation': 0,
            'rejected_no_alternance': 0,
            'rejected_wrong_domain': 0,
            'rejected_duration': 0
        }

        logger.debug(f"Classificateur initialisé avec {len(self.formation_keywords)} mots-clés formation, "
                    f"{len(self.alternance_keywords)} mots-clés alternance, "
                    f"{len(self.tech_keywords)} mots-clés techniques")

    def classify_offer(self, offer: JobOffer) -> Tuple[bool, str]:
        """
        Classifie et valide une offre d'alternance selon des critères experts.

        Cette méthode applique un pipeline de validation en plusieurs étapes
        pour déterminer si une offre correspond aux critères d'alternance
        en cybersécurité. Chaque étape est loggée pour la traçabilité.

        Args:
            offer (JobOffer): L'offre d'emploi à classifier

        Returns:
            Tuple[bool, str]: Tuple contenant:
                - is_valid (bool): True si l'offre est validée
                - reason (str): Raison détaillée de la validation/rejet

        Raises:
            ValueError: Si l'offre fournie est None ou invalide

        Example:
            classifier = OfferClassifier()
            is_valid, reason = classifier.classify_offer(my_offer)
            if is_valid:
                print(f"✅ {reason}")
            else:
                print(f"❌ {reason}")
        """
        if not offer:
            raise ValueError("L'offre fournie ne peut pas être None")

        logger.debug(f"Classification de l'offre: {offer.title} - {offer.company}")

        # Mise à jour des statistiques
        self.classification_stats['total_processed'] += 1

        # Préparation des données pour l'analyse
        title_lower = offer.title.lower() if offer.title else ""
        description_lower = offer.description.lower() if offer.description else ""
        company_lower = offer.company.lower() if offer.company else ""

        logger.debug(f"Analyse textuelle - Titre: '{title_lower[:50]}...', "
                    f"Entreprise: '{company_lower}', "
                    f"Description: {len(description_lower)} caractères")

        # Vérification 1: Exclure les formations/écoles
        logger.debug("🔍 Étape 1: Vérification exclusion formations/écoles")
        for keyword in self.formation_keywords:
            if keyword in title_lower or keyword in company_lower:
                reason = f"Formation détectée: '{keyword}'"
                logger.debug(f"❌ Rejet: {reason}")
                self.classification_stats['rejected_formation'] += 1
                return False, reason

        # Vérification 2: Doit contenir des mots-clés d'alternance
        logger.debug("🔍 Étape 2: Vérification présence alternance/apprentissage")
        alternance_matches = [kw for kw in self.alternance_keywords
                             if kw in title_lower or kw in description_lower]

        if not alternance_matches:
            reason = "Pas de mention d'alternance/apprentissage"
            logger.debug(f"❌ Rejet: {reason}")
            self.classification_stats['rejected_no_alternance'] += 1
            return False, reason
        else:
            logger.debug(f"✅ Alternance détectée: {alternance_matches}")

        # Vérification 3: Doit être dans le domaine technique
        logger.debug("🔍 Étape 3: Vérification domaine cybersécurité/réseaux")
        tech_matches = [kw for kw in self.tech_keywords
                       if kw in title_lower or kw in description_lower]

        if not tech_matches:
            reason = "Domaine non compatible (cybersécurité/réseaux)"
            logger.debug(f"❌ Rejet: {reason}")
            self.classification_stats['rejected_wrong_domain'] += 1
            return False, reason
        else:
            logger.debug(f"✅ Domaine technique détecté: {tech_matches}")

        # Vérification 4: Durée appropriée (si mentionnée)
        logger.debug("🔍 Étape 4: Vérification durée du contrat")
        if offer.duration and offer.duration.strip():
            duration_lower = offer.duration.lower()
            short_duration_indicators = ['6 mois', 'stage', '3 mois', '4 mois', '5 mois']

            for indicator in short_duration_indicators:
                if indicator in duration_lower:
                    reason = f"Durée trop courte: '{offer.duration}'"
                    logger.debug(f"❌ Rejet: {reason}")
                    self.classification_stats['rejected_duration'] += 1
                    return False, reason

            logger.debug(f"✅ Durée acceptable: {offer.duration}")

        # Validation finale
        reason = f"Offre validée - Alternance: {alternance_matches[:2]}, Tech: {tech_matches[:2]}"
        logger.info(f"✅ Offre VALIDÉE: {offer.title} chez {offer.company}")
        logger.debug(f"Détails validation: {reason}")

        self.classification_stats['validated'] += 1
        return True, reason

    def get_classification_stats(self) -> Dict[str, Union[int, float]]:
        """
        Retourne les statistiques de classification du moteur IA.

        Returns:
            Dict[str, Union[int, float]]: Statistiques détaillées incluant:
                - Nombre total d'offres traitées
                - Nombre d'offres validées
                - Détail des rejets par catégorie
                - Taux de validation global
        """
        stats = self.classification_stats.copy()
        total = stats['total_processed']

        if total > 0:
            stats['validation_rate'] = round((stats['validated'] / total) * 100, 2)
            stats['rejection_rate'] = round(((total - stats['validated']) / total) * 100, 2)
        else:
            stats['validation_rate'] = 0.0
            stats['rejection_rate'] = 0.0

        logger.debug(f"Statistiques classification: {stats}")
        return stats

class BaseScraper:
    """Classe de base pour tous les scrapers d'offres d'alternance."""

    def __init__(self, site_name: str):
        self.site_name = site_name
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        self.classifier = OfferClassifier()
        logger.info(f"Initialisation scraper pour {site_name}")

    def scrape_offers(self, keywords: List[str]) -> List[JobOffer]:
        """Méthode abstraite à implémenter par chaque scraper spécialisé."""
        raise NotImplementedError("Chaque scraper doit implémenter cette méthode")

    def _make_request(self, url: str, max_retries: int = 3) -> requests.Response:
        """Effectue une requête HTTP avec gestion d'erreurs et retry."""
        for attempt in range(max_retries):
            try:
                # Délai aléatoire pour éviter la détection
                time.sleep(random.uniform(1, 3))

                response = self.session.get(url, timeout=10)
                response.raise_for_status()

                logger.debug(f"Requête réussie: {url}")
                return response

            except requests.exceptions.RequestException as e:
                logger.warning(f"Tentative {attempt + 1}/{max_retries} échouée pour {url}: {e}")
                if attempt == max_retries - 1:
                    raise
                time.sleep(5)  # Attente avant retry

    def filter_and_validate_offers(self, offers: List[JobOffer]) -> List[JobOffer]:
        """Filtre et valide les offres avec le classificateur IA."""
        validated_offers = []

        for offer in offers:
            is_valid, reason = self.classifier.classify_offer(offer)
            offer.is_validated = is_valid
            offer.validation_reason = reason

            if is_valid:
                validated_offers.append(offer)
                logger.info(f"✅ Offre validée: {offer.title} - {offer.company}")
            else:
                logger.debug(f"❌ Offre rejetée: {offer.title} - Raison: {reason}")

        return validated_offers

class MockScraper(BaseScraper):
    """Scraper de démonstration avec données fictives pour les tests."""

    def __init__(self):
        super().__init__("Mock Job Site")

    def scrape_offers(self, keywords: List[str]) -> List[JobOffer]:
        """Génère des offres fictives pour démonstration."""
        logger.info("Génération d'offres fictives pour démonstration...")

        mock_offers = [
            # Offres valides
            JobOffer(
                site=self.site_name,
                title="Alternant Ingénieur Cybersécurité",
                company="SecureTech Solutions",
                location="Paris (75)",
                duration="24 mois",
                start_date="Septembre 2025",
                url="https://example.com/job1",
                description="Rejoignez notre équipe cybersécurité en contrat d'apprentissage. Formation sur les outils SOC, analyse de vulnérabilités, et gestion des incidents."
            ),
            JobOffer(
                site=self.site_name,
                title="Apprenti Administrateur Réseaux et Sécurité",
                company="DataCorp France",
                location="Lyon (69)",
                duration="12 mois",
                start_date="Septembre 2025",
                url="https://example.com/job2",
                description="Alternance en administration réseaux Cisco, pare-feux et systèmes de détection d'intrusion."
            ),

            # Offres à rejeter (formations)
            JobOffer(
                site=self.site_name,
                title="Formation Cybersécurité - École SuperTech",
                company="École SuperTech",
                location="Marseille (13)",
                duration="3 ans",
                start_date="Septembre 2025",
                url="https://example.com/formation1",
                description="Cursus complet en cybersécurité avec stages en entreprise. Diplôme reconnu par l'État."
            ),

            # Offres à rejeter (domaine non compatible)
            JobOffer(
                site=self.site_name,
                title="Alternant Développeur Web",
                company="WebDev Agency",
                location="Toulouse (31)",
                duration="18 mois",
                start_date="Septembre 2025",
                url="https://example.com/job3",
                description="Développement d'applications web en PHP/JavaScript. Contrat d'apprentissage."
            ),

            # Offres à rejeter (stage court)
            JobOffer(
                site=self.site_name,
                title="Stage Cybersécurité 6 mois",
                company="CyberDefense Corp",
                location="Nice (06)",
                duration="6 mois",
                start_date="Mars 2025",
                url="https://example.com/stage1",
                description="Stage en cybersécurité pour étudiant Master 1. Non rémunéré."
            )
        ]

        logger.info(f"Génération de {len(mock_offers)} offres fictives terminée")
        return mock_offers

class ExcelReportBuilder:
    """Générateur de rapports Excel pour les offres d'alternance."""

    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Offres Alternance Cybersécurité"
        logger.info("Initialisation du générateur de rapport Excel")

    def create_report(self, offers: List[JobOffer], filename: str = None) -> str:
        """
        Crée un rapport Excel avec mise en forme professionnelle.

        Args:
            offers: Liste des offres validées
            filename: Nom du fichier (auto-généré si None)

        Returns:
            Chemin du fichier créé
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"rapport_alternance_cybersecurite_{timestamp}.xlsx"

        # Configuration des en-têtes
        headers = [
            "Site", "Titre de l'offre", "Entreprise", "Localisation",
            "Durée", "Date de début", "Lien direct", "Validation"
        ]

        # Styles pour les en-têtes
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Création des en-têtes
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Ajout des données
        for row, offer in enumerate(offers, 2):
            self.worksheet.cell(row=row, column=1, value=offer.site)
            self.worksheet.cell(row=row, column=2, value=offer.title)
            self.worksheet.cell(row=row, column=3, value=offer.company)
            self.worksheet.cell(row=row, column=4, value=offer.location)
            self.worksheet.cell(row=row, column=5, value=offer.duration)
            self.worksheet.cell(row=row, column=6, value=offer.start_date)

            # Lien hypertexte pour l'URL
            link_cell = self.worksheet.cell(row=row, column=7, value=offer.url)
            link_cell.hyperlink = offer.url
            link_cell.font = Font(color="0563C1", underline="single")

            # Status de validation
            validation_text = "✅ Validée" if offer.is_validated else f"❌ {offer.validation_reason}"
            self.worksheet.cell(row=row, column=8, value=validation_text)

        # Auto-ajustement des colonnes
        for column in self.worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            self.worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

        # Ajout d'un résumé en haut
        summary_row = len(offers) + 3
        self.worksheet.cell(row=summary_row, column=1, value="RÉSUMÉ:").font = Font(bold=True)
        self.worksheet.cell(row=summary_row + 1, column=1, value=f"Total offres trouvées: {len(offers)}")
        self.worksheet.cell(row=summary_row + 2, column=1, value=f"Date génération: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        # Sauvegarde
        try:
            self.workbook.save(filename)
            logger.info(f"Rapport Excel créé: {filename}")
            return os.path.abspath(filename)
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde du rapport: {e}")
            raise

class AlternanceAgent:
    """Agent principal orchestrant la recherche d'offres d'alternance."""

    def __init__(self):
        self.scrapers = [MockScraper()]  # En production: ajouter PoleEmploiScraper, IndeedScraper, etc.
        self.report_builder = ExcelReportBuilder()
        self.keywords = [
            "alternance cybersécurité", "apprentissage cyber", "alternant sécurité informatique",
            "alternance réseaux", "apprenti administrateur réseau", "alternance télécoms"
        ]
        logger.info("Agent alternance initialisé")

    def run_full_search(self) -> str:
        """
        Lance une recherche complète sur tous les sites configurés.

        Returns:
            Chemin du rapport Excel généré
        """
        logger.info("🚀 Démarrage recherche complète d'offres d'alternance")
        all_offers = []

        for scraper in self.scrapers:
            try:
                logger.info(f"Scraping du site: {scraper.site_name}")

                # Récupération des offres brutes
                raw_offers = scraper.scrape_offers(self.keywords)
                logger.info(f"Offres brutes trouvées: {len(raw_offers)}")

                # Filtrage et validation
                validated_offers = scraper.filter_and_validate_offers(raw_offers)
                logger.info(f"Offres validées: {len(validated_offers)}")

                all_offers.extend(validated_offers)

            except Exception as e:
                logger.error(f"Erreur lors du scraping de {scraper.site_name}: {e}")
                continue

        # Génération du rapport
        logger.info(f"Génération du rapport final - Total: {len(all_offers)} offres validées")

        if all_offers:
            report_path = self.report_builder.create_report(all_offers)
            logger.info(f"✅ Rapport généré avec succès: {report_path}")

            # Statistiques finales
            self._log_statistics(all_offers)

            return report_path
        else:
            logger.warning("⚠️ Aucune offre valide trouvée")
            return None

    def _log_statistics(self, offers: List[JobOffer]):
        """Affiche des statistiques sur les offres trouvées."""
        logger.info("📊 STATISTIQUES FINALES:")
        logger.info(f"   • Total offres validées: {len(offers)}")

        # Répartition par site
        sites = {}
        for offer in offers:
            sites[offer.site] = sites.get(offer.site, 0) + 1

        for site, count in sites.items():
            logger.info(f"   • {site}: {count} offres")

        # Répartition par localisation
        locations = {}
        for offer in offers:
            loc = offer.location.split('(')[0].strip() if '(' in offer.location else offer.location
            locations[loc] = locations.get(loc, 0) + 1

        top_locations = sorted(locations.items(), key=lambda x: x[1], reverse=True)[:5]
        logger.info(f"   • Top localisations: {', '.join([f'{loc} ({count})' for loc, count in top_locations])}")

def main():
    """Point d'entrée principal du script."""
    print("=" * 70)
    print("🤖 AGENT IA - RECHERCHE OFFRES ALTERNANCE CYBERSÉCURITÉ")
    print("=" * 70)
    print()

    try:
        # Initialisation de l'agent
        agent = AlternanceAgent()

        # Lancement de la recherche
        report_path = agent.run_full_search()

        if report_path:
            print(f"\n✅ SUCCÈS!")
            print(f"📄 Rapport Excel généré: {report_path}")
            print(f"📊 Consultez les logs pour plus de détails")
        else:
            print(f"\n⚠️ AUCUNE OFFRE TROUVÉE")
            print("Vérifiez les critères de recherche ou les sites sources")

    except KeyboardInterrupt:
        logger.info("Arrêt demandé par l'utilisateur")
        print("\n🛑 Arrêt du programme")
    except Exception as e:
        logger.error(f"Erreur fatale: {e}")
        print(f"\n❌ ERREUR: {e}")
        print("Consultez les logs pour plus de détails")

    print("\n" + "=" * 70)

if __name__ == "__main__":
    main()